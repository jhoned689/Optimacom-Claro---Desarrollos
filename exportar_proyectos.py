# -*- coding: utf-8 -*-
"""Exporta cada desarrollo como proyecto con sus actividades agrupadas."""
from __future__ import annotations

import csv
import json
import re
import shutil
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).parent
PROYECTOS_DIR = ROOT / "proyectos"
XLSX = ROOT / "PBI Desarrollos.xlsx"
JSON_OUT = ROOT / "data" / "proyectos.json"

ACTIVITY_HEADERS = ["Fase", "Actividades", "Entregable", "Fecha", "Estado", "% Avance", "Notas"]

# Nombre en ejecución → nombre en DESARROLLOS LISTOS (avance PBI)
AVANCE_PBI_MAP: dict[str, str] = {
    "REVISION ACTA PACC - FACC SERVICIOS": "AUTOMATIZACION AUDITORIA ACTAS PACC Y FACC SERV / LIC / TSS",
    "REVISION ACTA PACC - FACC HW": "AUTOMATIZACION AUDITORIA ACTAS PACC Y FACC HARDWARE",
    "REVISION ACTAS LIQUIDACION - HW": "AUTOMATIZACION AUDITORIA ACTAS PACC Y FACC HARDWARE",
    "VALIDACION HW LINEA A LINEA NOKIA-LOCAL": "AUTOMATIZACION CHECK HW MEDIANTE IA",
    "REVISION LINEA A LINEA HW": "AUTOMATIZACION CHECK HW MEDIANTE IA",
    "APP PENALIDADES": "APP GESTION PENALIDADES INT / ONAIR / HALLAZGOS",
    "RADICACION ACTAS DE LIQUIDACION (FIRMAS)": "GESTION DE ACTAS FIRMAS DE LIQUIDACION",
    "GESTION FIRMAS ACTAS DE LIQUIDACION": "GESTION DE ACTAS FIRMAS DE LIQUIDACION",
    "ACTA SISTEMATIZADA / CASOS DE BAJA": "ACTAS SISTEMATIZADAS / CASOS DE RECOLECCION - LOCAL",
    "CREACION ACTA SISTEMATIZADA HW BAJA": "ACTAS SISTEMATIZADAS / CASOS DE RECOLECCION - LOCAL",
    "CREACION CASO DE RECOLECCION HW BAJA": "ACTAS SISTEMATIZADAS / CASOS DE RECOLECCION - LOCAL",
    "APP CHECK HW": "CHECK HW",
    "APP CHECK DOC": "CHECK DOCUMENTACION",
    "BI CONTRATOS": "PBI GESTION DE CONTRATOS",
    "PBI ACTAS DE BAJA": "PBI ACTAS DE BAJA NOKIA",
    "BITACORAS DIRECTOS": "BITACORAS DIRECTOS",
    "APP GESTION DE ACCESOS ESCALADOS": "GESTION DE ACCESOS ESCALADOS",
    "GESTION ACCESOS EB": "GESTION DE ACCESOS ESCALADOS",
    "SOLICITUD ACCESO EB": "GESTION DE ACCESOS ESCALADOS",
    "ONAIR": "AUTOMATIZACION REPORTES RI",
}

ESTADO_ORDEN = {"HOY": 0, "PROCESO": 1, "PENDIENTE": 2, "REALIZADO": 3}


def slugify(nombre: str) -> str:
    s = nombre.lower().strip()
    s = re.sub(r"[^\w\s-]", "", s, flags=re.UNICODE)
    s = re.sub(r"[\s_]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s[:80] or "proyecto"


def fmt_fecha(val: Any) -> str:
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    if val in (None, "", "PEND"):
        return "—"
    return str(val)


def fmt_pct(val: Any) -> str:
    if val is None or val == "":
        return "—"
    if isinstance(val, (int, float)):
        if val <= 1:
            return f"{int(round(val * 100))}%"
        return f"{int(val)}%"
    return str(val)


def parse_fecha(val: Any) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def leer_avance_pbi(wb: openpyxl.Workbook) -> dict[str, float]:
    ws = wb["DESARROLLOS LISTOS"]
    out: dict[str, float] = {}
    for r in range(3, ws.max_row + 1):
        nombre = ws.cell(r, 2).value
        avance = ws.cell(r, 3).value
        if nombre and avance is not None:
            out[str(nombre)] = float(avance)
    return out


def leer_actividades(wb: openpyxl.Workbook) -> dict[str, list[dict[str, Any]]]:
    ws = wb["EJECUCION DESAROLLO"]
    proyectos: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        nombre = ws.cell(r, 1).value
        if not nombre:
            continue
        act = {
            "fase": ws.cell(r, 2).value or "",
            "actividades": ws.cell(r, 3).value or "",
            "entregable": ws.cell(r, 4).value or "",
            "fecha": ws.cell(r, 5).value,
            "estado": ws.cell(r, 6).value or "",
            "avance": ws.cell(r, 7).value,
            "notas": ws.cell(r, 8).value or "",
        }
        proyectos[str(nombre)].append(act)
    return dict(proyectos)


def resumen_proyecto(nombre: str, actividades: list[dict], avance_pbi: dict[str, float]) -> dict[str, Any]:
    estados = [a["estado"] for a in actividades]
    realizadas = sum(1 for e in estados if e == "REALIZADO")
    en_curso = sum(1 for e in estados if e in ("HOY", "PROCESO"))
    pendientes = sum(1 for e in estados if e == "PENDIENTE")

    if any(e == "HOY" for e in estados):
        estado_actual = "HOY"
    elif any(e == "PROCESO" for e in estados):
        estado_actual = "PROCESO"
    elif pendientes and realizadas:
        estado_actual = "EN CURSO"
    elif realizadas == len(actividades):
        estado_actual = "REALIZADO"
    else:
        estado_actual = "PENDIENTE"

    avances_num = [a["avance"] for a in actividades if isinstance(a["avance"], (int, float))]
    avance_calc = max(avances_num) if avances_num else realizadas / len(actividades)

    pbi_key = AVANCE_PBI_MAP.get(nombre)
    if pbi_key and pbi_key in avance_pbi:
        avance = avance_pbi[pbi_key]
    else:
        avance = avance_calc

    futuras = [
        a for a in actividades
        if a["estado"] in ("HOY", "PROCESO", "PENDIENTE") and parse_fecha(a["fecha"])
    ]
    futuras.sort(key=lambda a: parse_fecha(a["fecha"]) or date.max)
    proxima = futuras[0] if futuras else None

    fechas = [parse_fecha(a["fecha"]) for a in actividades if parse_fecha(a["fecha"])]
    inicio = min(fechas) if fechas else None

    return {
        "nombre": nombre,
        "slug": slugify(nombre),
        "estado": estado_actual,
        "avance": avance,
        "total_actividades": len(actividades),
        "realizadas": realizadas,
        "en_curso": en_curso,
        "pendientes": pendientes,
        "inicio": fmt_fecha(inicio) if inicio else "—",
        "proxima_fecha": fmt_fecha(proxima["fecha"]) if proxima else "—",
        "proxima_entregable": proxima["entregable"] if proxima else "—",
        "actividades": actividades,
    }


def escribir_actividades_csv(path: Path, actividades: list[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(ACTIVITY_HEADERS)
        for a in actividades:
            w.writerow([
                a["fase"],
                a["actividades"],
                a["entregable"],
                fmt_fecha(a["fecha"]),
                a["estado"],
                a["avance"] if a["avance"] is not None else "",
                a["notas"],
            ])


def escribir_proyecto_readme(path: Path, p: dict) -> None:
    lines = [
        f"# {p['nombre']}",
        "",
        "| Campo | Valor |",
        "|-------|-------|",
        f"| **Estado** | {p['estado']} |",
        f"| **Avance** | {fmt_pct(p['avance'])} |",
        f"| **Actividades** | {p['realizadas']} realizadas · {p['en_curso']} en curso · {p['pendientes']} pendientes |",
        f"| **Inicio** | {p['inicio']} |",
        f"| **Próxima entrega** | {p['proxima_fecha']} |",
        f"| **Próximo entregable** | {p['proxima_entregable']} |",
        "",
        "## Actividades",
        "",
        "| Fase | Actividad | Entregable | Fecha | Estado | % |",
        "|------|-----------|------------|-------|--------|---|",
    ]
    for a in p["actividades"]:
        lines.append(
            f"| {a['fase']} | {a['actividades']} | {a['entregable']} | "
            f"{fmt_fecha(a['fecha'])} | {a['estado']} | {fmt_pct(a['avance'])} |"
        )
    lines.extend(["", "[← Volver al índice de proyectos](../README.md)", ""])
    path.write_text("\n".join(lines), encoding="utf-8")


def escribir_indice(proyectos: list[dict]) -> None:
    activos = [p for p in proyectos if p["estado"] in ("HOY", "PROCESO", "EN CURSO")]
    pendientes = [p for p in proyectos if p["estado"] == "PENDIENTE"]
    realizados = [p for p in proyectos if p["estado"] == "REALIZADO"]

    def tabla(items: list[dict]) -> list[str]:
        rows = [
            "| Proyecto | Avance | Estado | Actividades | Próxima fecha |",
            "|----------|--------|--------|-------------|---------------|",
        ]
        for p in items:
            rows.append(
                f"| [{p['nombre']}]({p['slug']}/README.md) | {fmt_pct(p['avance'])} | "
                f"{p['estado']} | {p['total_actividades']} | {p['proxima_fecha']} |"
            )
        return rows

    content = [
        "# Proyectos de desarrollo",
        "",
        "Cada carpeta es un **proyecto (desarrollo)** con sus actividades dentro.",
        "",
        f"**Total:** {len(proyectos)} proyectos · "
        f"{sum(p['total_actividades'] for p in proyectos)} actividades",
        "",
        "## En curso",
        "",
        *tabla(sorted(activos, key=lambda p: ESTADO_ORDEN.get(p["estado"], 9))),
        "",
        "## Pendientes",
        "",
        *tabla(sorted(pendientes, key=lambda p: p["nombre"])),
        "",
        "## Realizados",
        "",
        *tabla(sorted(realizados, key=lambda p: p["nombre"])),
        "",
        "## Estructura",
        "",
        "```",
        "proyectos/",
        "  nombre-del-proyecto/",
        "    README.md       ← ficha del proyecto + tabla de actividades",
        "    actividades.csv ← detalle exportado",
        "```",
        "",
    ]
    (PROYECTOS_DIR / "README.md").write_text("\n".join(content), encoding="utf-8")


def actualizar_hoja_proyectos_excel(wb: openpyxl.Workbook, proyectos: list[dict]) -> None:
    nombre_hoja = "PROYECTOS"
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]
    ws = wb.create_sheet(nombre_hoja)
    headers = [
        "Proyecto", "Slug", "% Avance", "Estado", "Total actividades",
        "Realizadas", "En curso", "Pendientes", "Inicio", "Próxima fecha", "Próximo entregable",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for i, p in enumerate(sorted(proyectos, key=lambda x: (ESTADO_ORDEN.get(x["estado"], 9), x["nombre"])), 2):
        ws.cell(i, 1, p["nombre"])
        ws.cell(i, 2, p["slug"])
        ws.cell(i, 3, p["avance"])
        ws.cell(i, 4, p["estado"])
        ws.cell(i, 5, p["total_actividades"])
        ws.cell(i, 6, p["realizadas"])
        ws.cell(i, 7, p["en_curso"])
        ws.cell(i, 8, p["pendientes"])
        ws.cell(i, 9, p["inicio"])
        ws.cell(i, 10, p["proxima_fecha"])
        ws.cell(i, 11, p["proxima_entregable"])


def exportar_proyectos(actualizar_excel: bool = True) -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    avance_pbi = leer_avance_pbi(wb)
    agrupado = leer_actividades(wb)

    proyectos = [
        resumen_proyecto(nombre, acts, avance_pbi)
        for nombre, acts in agrupado.items()
    ]
    proyectos.sort(key=lambda p: (ESTADO_ORDEN.get(p["estado"], 9), -p["avance"], p["nombre"]))

    if PROYECTOS_DIR.exists():
        for item in PROYECTOS_DIR.iterdir():
            if item.is_dir():
                shutil.rmtree(item)
    PROYECTOS_DIR.mkdir(exist_ok=True)
    ROOT.joinpath("data").mkdir(exist_ok=True)

    json_data = []
    for p in proyectos:
        carpeta = PROYECTOS_DIR / p["slug"]
        carpeta.mkdir(parents=True, exist_ok=True)
        escribir_proyecto_readme(carpeta / "README.md", p)
        escribir_actividades_csv(carpeta / "actividades.csv", p["actividades"])
        item = {k: v for k, v in p.items() if k != "actividades"}
        item["actividades"] = [
            {**a, "fecha": fmt_fecha(a["fecha"])} for a in p["actividades"]
        ]
        json_data.append(item)

    escribir_indice(proyectos)
    JSON_OUT.write_text(json.dumps(json_data, ensure_ascii=False, indent=2), encoding="utf-8")

    if actualizar_excel:
        wb_rw = openpyxl.load_workbook(XLSX)
        actualizar_hoja_proyectos_excel(wb_rw, proyectos)
        wb_rw.save(XLSX)

    return proyectos


if __name__ == "__main__":
    ps = exportar_proyectos()
    print(f"Proyectos exportados: {len(ps)}")
    print(f"Carpeta: {PROYECTOS_DIR}")

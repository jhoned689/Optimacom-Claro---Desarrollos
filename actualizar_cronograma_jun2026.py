# -*- coding: utf-8 -*-
"""
Actualiza PBI Desarrollos.xlsx y exporta CSV para GitHub.

Regla de fechas: entregas solo miércoles y viernes.
"""
from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
XLSX = ROOT / "PBI Desarrollos.xlsx"
CSV_EJECUCION = ROOT / "data" / "ejecucion-desarrollo.csv"
CSV_AVANCE = ROOT / "data" / "avance-proyectos.csv"

# Miércoles y viernes (año, mes, día)
D = {
    "hoy": date(2026, 6, 24),       # mié
    "vie_27": date(2026, 6, 27),
    "mie_01": date(2026, 7, 1),
    "vie_03": date(2026, 7, 3),
    "mie_08": date(2026, 7, 8),
    "vie_10": date(2026, 7, 10),
    "mie_15": date(2026, 7, 15),
    "vie_17": date(2026, 7, 17),
    "mie_22": date(2026, 7, 22),
    "vie_24": date(2026, 7, 24),
    "mie_29": date(2026, 7, 29),
    "vie_31": date(2026, 7, 31),
    "mie_05_ago": date(2026, 8, 5),
    "vie_07_ago": date(2026, 8, 7),
    "mie_12_ago": date(2026, 8, 12),
    "vie_14_ago": date(2026, 8, 14),
    "mie_19_ago": date(2026, 8, 19),
    "vie_21_ago": date(2026, 8, 21),
    "mie_26_ago": date(2026, 8, 26),
    "vie_28_ago": date(2026, 8, 28),
    "mie_02_sep": date(2026, 9, 2),
    "vie_04_sep": date(2026, 9, 4),
    "mie_09_sep": date(2026, 9, 9),
    "vie_11_sep": date(2026, 9, 11),
    "mie_16_sep": date(2026, 9, 16),
    "vie_18_sep": date(2026, 9, 18),
    "mie_23_sep": date(2026, 9, 23),
    "vie_25_sep": date(2026, 9, 25),
    "mie_30_sep": date(2026, 9, 30),
    "vie_02_oct": date(2026, 10, 2),
    "mie_07_oct": date(2026, 10, 7),
    "vie_09_oct": date(2026, 10, 9),
    "mie_14_oct": date(2026, 10, 14),
    "vie_16_oct": date(2026, 10, 16),
}


def dt(d: date) -> datetime:
    return datetime(d.year, d.month, d.day)


# DESARROLLO, Fase, Actividades, Entregable, Fecha, Estado, % Avance, Notas
NUEVAS_FILAS = [
    (
        "REVISION ACTA PACC - FACC SERVICIOS",
        "Pruebas",
        "Socialización inicio pruebas usuarios",
        "Módulo servicios listo para prueba con actas reales",
        dt(D["hoy"]),
        "HOY",
        0.85,
        "Mié 24/06 — flujo V7 + validaciones core",
    ),
    (
        "REVISION ACTA PACC - FACC HW",
        "Pruebas",
        "Socialización inicio pruebas usuarios",
        "Módulo HW demostración funcional parcial",
        dt(D["hoy"]),
        "HOY",
        0.5,
        "Mié 24/06 — alcance parcial vs servicios",
    ),
    (
        "REVISION ACTAS LIQUIDACION - HW",
        "Pruebas",
        "Socialización módulo auditoría HW",
        "Primera demostración checklist HW",
        dt(D["hoy"]),
        "HOY",
        0.5,
        "Mié 24/06 — alineado con PACC-FACC HW",
    ),
    (
        "VALIDACION HW LINEA A LINEA NOKIA-LOCAL",
        "Levantamiento",
        "Presentación modelo automatización videos IA",
        "Demo investigación revisión videos HW con IA",
        dt(D["hoy"]),
        "HOY",
        0.15,
        "Mié 24/06 — modelo/concepto, no productivo",
    ),
    (
        "REVISION LINEA A LINEA HW",
        "Investigación",
        "Análisis videos proceso Maira",
        "Documento oportunidades automatización video",
        dt(D["hoy"]),
        "HOY",
        0.2,
        "Mié 24/06 — base validación línea a línea",
    ),
    (
        "REVISION ACTA PACC - FACC SERVICIOS",
        "Cambios",
        "Ajustes post-prueba desarrollo",
        "Segunda versión con correcciones servicios",
        dt(D["vie_27"]),
        "PENDIENTE",
        0.85,
        "Vie 27/06 — ajustes + pruebas servicios",
    ),
    (
        "REVISION ACTA PACC - FACC HW",
        "Cambios",
        "Ajustes post-prueba desarrollo",
        "Segunda versión con correcciones HW",
        dt(D["vie_27"]),
        "PENDIENTE",
        0.55,
        "Vie 27/06 — ajustes + pruebas HW",
    ),
    (
        "REVISION ACTAS LIQUIDACION - HW",
        "Cambios",
        "Ajustes pruebas usuarios",
        "Segunda versión funcional estable HW",
        dt(D["vie_27"]),
        "PENDIENTE",
        0.55,
        "Vie 27/06 — ajustes módulo HW",
    ),
    (
        "REVISION ACTA PACC - FACC SERVICIOS",
        "Pruebas",
        "Segunda ronda validación usuarios",
        "Validación operativa servicios post-ajustes",
        dt(D["vie_27"]),
        "PENDIENTE",
        0.9,
        "Vie 27/06 — meta ~90% servicios",
    ),
    (
        "REVISION ACTA PACC - FACC HW",
        "Pruebas",
        "Segunda ronda validación usuarios",
        "Validación operativa HW post-ajustes",
        dt(D["vie_27"]),
        "PENDIENTE",
        0.6,
        "Vie 27/06 — pruebas HW",
    ),
    (
        "REVISION ACTA PACC - FACC SERVICIOS",
        "Pruebas",
        "Estabilización y cierre hallazgos",
        "Servicios listos para documentación",
        dt(D["mie_01"]),
        "PENDIENTE",
        0.92,
        "Mié 01/07",
    ),
    (
        "REVISION ACTA PACC - FACC HW",
        "Pruebas",
        "Estabilización módulo HW",
        "HW estable para documentación",
        dt(D["mie_01"]),
        "PENDIENTE",
        0.65,
        "Mié 01/07",
    ),
    (
        "REVISION ACTA PACC - FACC SERVICIOS",
        "Documentación",
        "Manual técnico y usuario",
        "Documentación entregada servicios",
        dt(D["vie_03"]),
        "PENDIENTE",
        0,
        "Vie 03/07",
    ),
    (
        "VALIDACION HW LINEA A LINEA NOKIA-LOCAL",
        "Diseño",
        "Diseño validación automatizada HW",
        "Modelo funcional validación línea a línea",
        dt(D["vie_03"]),
        "PENDIENTE",
        0.15,
        "Vie 03/07 — retomar post PACC/FACC",
    ),
    (
        "REVISION ACTA PACC - FACC HW",
        "Documentación",
        "Manual técnico y usuario",
        "Documentación entregada HW",
        dt(D["mie_08"]),
        "PENDIENTE",
        0,
        "Mié 08/07",
    ),
    (
        "REVISION ACTAS LIQUIDACION - HW",
        "Lanzamiento",
        "Socialización módulo liberado",
        "Módulo HW liberado usuarios piloto",
        dt(D["vie_10"]),
        "PENDIENTE",
        0,
        "Vie 10/07 — meta julio cronograma Oscar",
    ),
    (
        "VALIDACION HW LINEA A LINEA NOKIA-LOCAL",
        "Desarrollo",
        "Construcción validación HW piloto",
        "Validación automática funcional piloto",
        dt(D["mie_15"]),
        "PENDIENTE",
        0,
        "Mié 15/07",
    ),
    (
        "GESTION FIRMAS ACTAS DE LIQUIDACION",
        "Producción",
        "Salida productiva",
        "Entrada a producción tras aval Gestión de Riesgos",
        dt(D["mie_15"]),
        "PROCESO",
        0.9,
        "Mié 15/07 — app ~90%, pendiente aval producción",
    ),
    (
        "RADICACION ACTAS DE LIQUIDACION (FIRMAS)",
        "Producción",
        "Estabilización y soporte",
        "Versión estable en producción",
        dt(D["vie_17"]),
        "PROCESO",
        0.9,
        "Vie 17/07 — mismo paquete firmas, post-aval",
    ),
    (
        "RADICACION ACTAS DE LIQUIDACION (FIRMAS)",
        "Firmas Digitales",
        "Investigar Adobe Sign",
        "Propuesta firmas digitales (fase 2)",
        dt(D["mie_22"]),
        "PROCESO",
        0.1,
        "Mié 22/07 — baja prioridad, post-producción",
    ),
    (
        "CREACION CASO DE RECOLECCION HW BAJA",
        "Desarrollo",
        "Automatización creación casos",
        "Flujo funcional generación casos",
        dt(D["vie_07_ago"]),
        "PENDIENTE",
        0.3,
        "Vie 07/08 — Livan",
    ),
    (
        "CREACION ACTA SISTEMATIZADA HW BAJA",
        "Producción",
        "Estabilización y pruebas aliados",
        "Acta HW baja operativa",
        dt(D["mie_12_ago"]),
        "PENDIENTE",
        0.7,
        "Mié 12/08 — Livan",
    ),
    (
        "GESTION ACCESOS EB",
        "Desarrollo",
        "Construcción automatización accesos",
        "Aplicación en pruebas funcional",
        dt(D["vie_14_ago"]),
        "PROCESO",
        0.4,
        "Vie 14/08 — Andrea",
    ),
    (
        "SOLICITUD ACCESO EB",
        "Desarrollo",
        "Construcción flujo solicitud",
        "Flujo funcional solicitud accesos",
        dt(D["mie_19_ago"]),
        "PENDIENTE",
        0.25,
        "Mié 19/08 — Andrea",
    ),
    (
        "REVISION FORMATOS FSE",
        "Levantamiento",
        "Revisión criterios validación FSE",
        "Documento validaciones FSE",
        dt(D["mie_02_sep"]),
        "PENDIENTE",
        0,
        "Mié 02/09 — cronograma Oscar",
    ),
    (
        "REVISION Y CONCEPTO DE TSS",
        "Levantamiento",
        "Revisión criterios validación TSS",
        "Documento validaciones TSS",
        dt(D["mie_02_sep"]),
        "PENDIENTE",
        0,
        "Mié 02/09 — Regionales Acceso",
    ),
    (
        "ACTUALIZACION POR SITIO EN PORTAL RF",
        "Levantamiento",
        "Revisión actualización RF",
        "Flujo funcional actualización portal RF",
        dt(D["vie_04_sep"]),
        "PENDIENTE",
        0,
        "Vie 04/09 — Jose Cardenas",
    ),
    (
        "REVISION FORMATOS FSE",
        "Diseño",
        "Diseño validaciones formatos",
        "Modelo automatización formatos FSE",
        dt(D["mie_09_sep"]),
        "PENDIENTE",
        0,
        "Mié 09/09",
    ),
    (
        "REVISION DOCUMENTACION DESINSTALACION / DESMONTE",
        "Desarrollo",
        "Construcción validaciones automáticas",
        "Primera versión funcional desmonte",
        dt(D["vie_11_sep"]),
        "PENDIENTE",
        0.1,
        "Vie 11/09 — solo levantamiento hecho",
    ),
    (
        "REVISION DOCUMENTACION INSTALACION / MIGRACION",
        "Desarrollo",
        "Construcción validaciones automáticas",
        "Primera versión funcional instalación",
        dt(D["vie_11_sep"]),
        "PENDIENTE",
        0.1,
        "Vie 11/09 — solo levantamiento hecho",
    ),
    (
        "ASIGNACION ACTIVIDAD ALIADO",
        "Diseño",
        "Diseño flujo asignación actividades",
        "Primera versión automatización asignación",
        dt(D["mie_16_sep"]),
        "PENDIENTE",
        0,
        "Mié 16/09",
    ),
    (
        "APROBACION ID TRABAJOS SITE ACCESS",
        "Diseño",
        "Diseño aprobación IDs",
        "Flujo automatizado aprobaciones",
        dt(D["mie_16_sep"]),
        "PENDIENTE",
        0,
        "Mié 16/09",
    ),
    (
        "CREACION ID TRABAJOS SITE ACCESS",
        "Diseño",
        "Diseño creación IDs",
        "Flujo automatizado creación IDs",
        dt(D["vie_18_sep"]),
        "PENDIENTE",
        0,
        "Vie 18/09",
    ),
    (
        "REVISION SITE FOLDER NOKIA-LOCAL",
        "Desarrollo",
        "Construcción validación Site Folder",
        "Automatización revisión documental",
        dt(D["mie_23_sep"]),
        "PENDIENTE",
        0.05,
        "Mié 23/09",
    ),
    (
        "IMPLEMENTACION RUTA TX",
        "Levantamiento",
        "Revisión proceso implementación TX",
        "Documento flujo operativo TX",
        dt(D["mie_23_sep"]),
        "PENDIENTE",
        0,
        "Mié 23/09",
    ),
    (
        "PLANIFICACION MIGRACION / RUTA TX",
        "Levantamiento",
        "Identificación flujo migración",
        "Documento planificación automatizable",
        dt(D["vie_25_sep"]),
        "PENDIENTE",
        0,
        "Vie 25/09",
    ),
    (
        "CREACION USUARIO",
        "Diseño",
        "Diseño automatización creación usuarios",
        "Flujo automatizado creación usuario",
        dt(D["vie_25_sep"]),
        "PENDIENTE",
        0,
        "Vie 25/09 — Jessica Moreno",
    ),
    (
        "DESBLOQUEO USUARIO",
        "Diseño",
        "Diseño automatización desbloqueos",
        "Flujo automatizado desbloqueo",
        dt(D["mie_30_sep"]),
        "PENDIENTE",
        0,
        "Mié 30/09 — Jessica Moreno",
    ),
    (
        "GESTION DE OTS",
        "Diseño",
        "Diseño automatización OTS",
        "Primera estructura automatización OTS",
        dt(D["mie_30_sep"]),
        "PENDIENTE",
        0,
        "Mié 30/09",
    ),
    (
        "GESTION PERMISOS ACCESO / VENTANAS MANTENIMIENTO",
        "Diseño",
        "Diseño flujo permisos mantenimiento",
        "Flujo automatizado permisos",
        dt(D["vie_02_oct"]),
        "PENDIENTE",
        0,
        "Vie 02/10",
    ),
    (
        "SEGUIMIENTO IMPLEMENTACION TX",
        "Diseño",
        "Diseño seguimiento automatizado",
        "Modelo seguimiento implementación",
        dt(D["mie_07_oct"]),
        "PENDIENTE",
        0,
        "Mié 07/10",
    ),
    (
        "SEGUIMIENTO IMPLEMENTACION ACCESO",
        "Diseño",
        "Diseño seguimiento accesos",
        "Flujo seguimiento accesos",
        dt(D["mie_07_oct"]),
        "PENDIENTE",
        0,
        "Mié 07/10",
    ),
    (
        "SEGUIMIENTO ENTREGA HITOS INSTALACION E INTEGRACION",
        "Diseño",
        "Diseño seguimiento hitos",
        "Modelo control hitos integración",
        dt(D["vie_09_oct"]),
        "PENDIENTE",
        0,
        "Vie 09/10",
    ),
    (
        "SEGUIMIENTO RFIC",
        "Diseño",
        "Diseño seguimiento RFIC",
        "Flujo control RFIC",
        dt(D["vie_09_oct"]),
        "PENDIENTE",
        0,
        "Vie 09/10",
    ),
]

AVANCE_DESARROLLOS = {
    "AUTOMATIZACION AUDITORIA ACTAS PACC Y FACC SERV / LIC / TSS": 0.85,
    "AUTOMATIZACION AUDITORIA ACTAS PACC Y FACC HARDWARE": 0.5,
    "AUTOMATIZACION CHECK HW MEDIANTE IA": 0.15,
    "GESTION DE ACTAS FIRMAS DE LIQUIDACION": 0.9,
}

HEADERS = [
    "DESARROLLO",
    "Fase",
    "Actividades",
    "Entregable",
    "Fecha",
    "Estado",
    "% Avance",
    "Notas",
]


def actualizar_excel(start: int = 95) -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["EJECUCION DESAROLLO"]

    for c, h in enumerate(HEADERS, start=1):
        ws.cell(1, c, h)

    old_end = ws.max_row
    if old_end > start:
        ws.delete_rows(start, old_end - start + 1)

    for i, fila in enumerate(NUEVAS_FILAS):
        r = start + i
        for c, val in enumerate(fila, start=1):
            ws.cell(r, c, val)

    ws2 = wb["DESARROLLOS LISTOS"]
    for r in range(3, ws2.max_row + 1):
        nombre = ws2.cell(r, 2).value
        if nombre in AVANCE_DESARROLLOS:
            ws2.cell(r, 3, AVANCE_DESARROLLOS[nombre])

    wb.save(XLSX)
    return start + len(NUEVAS_FILAS) - 1


def _fmt_fecha(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    return str(val) if val is not None else ""


def exportar_csv() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ROOT.joinpath("data").mkdir(exist_ok=True)

    ws = wb["EJECUCION DESAROLLO"]
    with CSV_EJECUCION.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(HEADERS)
        for r in range(2, ws.max_row + 1):
            row = []
            for c in range(1, len(HEADERS) + 1):
                val = ws.cell(r, c).value
                if c == 5:
                    val = _fmt_fecha(val)
                row.append(val)
            if any(row):
                w.writerow(row)

    ws2 = wb["DESARROLLOS LISTOS"]
    with CSV_AVANCE.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["DESARROLLO", "% AVANCE", "INICIO", "FINALIZACION"])
        for r in range(3, ws2.max_row + 1):
            nombre = ws2.cell(r, 2).value
            if not nombre:
                continue
            w.writerow([
                nombre,
                ws2.cell(r, 3).value,
                _fmt_fecha(ws2.cell(r, 4).value),
                _fmt_fecha(ws2.cell(r, 5).value) if ws2.cell(r, 5).value != "--" else "--",
            ])


def main() -> None:
    end = actualizar_excel()
    exportar_csv()
    print(f"Excel: {XLSX}")
    print(f"Filas actualizadas: 95 - {end}")
    print(f"CSV: {CSV_EJECUCION}")
    print(f"CSV: {CSV_AVANCE}")


if __name__ == "__main__":
    main()
